import os
import sys
import json
import threading
import webbrowser
import traceback
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs
import pandas as pd

from analyse import (
    laad_rapport, bereken_dienstverdeling,
    bereken_senioren_overzicht, vergelijk_rapporten
)
from rooster import laad_medewerkers, laad_dienst_normen, analyseer_rooster
from werkplek import bouw_maand_werkplek
from maandoverzicht import bouw_maandoverzicht
from traject import analyseer_traject, samenvatting, genereer_inwerkschema, TRAJECTEN, laad_traject_data, sla_traject_data_op
from traject_optimizer import (laad_actieve_trajecten, sla_actieve_trajecten_op,
                                analyseer_traject_kansen, signaleer_vrijgekomen_plekken)
from analyse import laad_rapport, bereken_dienstverdeling, bereken_senioren_overzicht, vergelijk_rapporten, laad_uitsluitingen

# ── State ──────────────────────────────────────────────────────────────────────
rapporten      = {}
_medewerkers   = {}
_uitsluitingen = set()
_uitsluitingen = set()
_dienst_normen = {}
_rooster_cache = {}

def get_template(naam):
    base = os.path.dirname(os.path.abspath(__file__))
    pad = os.path.join(base, 'templates', naam)
    with open(pad, 'r', encoding='utf-8') as f:
        return f.read()

class Handler(BaseHTTPRequestHandler):

    def log_message(self, format, *args):
        pass

    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, html):
        body = html.encode('utf-8')
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        path   = parsed.path
        try:
            if path in ('/', '/index.html'):
                self.send_html(get_template('index.html'))
            elif path == '/api/rapporten':
                self.send_json({'rapporten': list(rapporten.keys())})
            elif path == '/api/update_check':
                import os as _os
                update_pad = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'update_beschikbaar.json')
                if _os.path.exists(update_pad):
                    with open(update_pad, 'r', encoding='utf-8') as f:
                        self.send_json({'update': True, **json.load(f)})
                else:
                    import json as _json2
                    versie_pad = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'version.json')
                    try:
                        with open(versie_pad, 'r', encoding='utf-8') as f:
                            v = _json2.load(f)
                    except:
                        v = {'versie': '?'}
                    self.send_json({'update': False, 'versie': v.get('versie','?')})
            elif path == '/api/vloer':
                qs = parse_qs(parsed.query)
                datum_str = qs.get('datum', [None])[0]
                if not datum_str or _huidig_rapport is None:
                    self.send_json({'diensten': []}); return
                try:
                    import re as _re
                    from datetime import date as _date
                    dag_dt = _date.fromisoformat(datum_str)
                    dag_df = _huidig_rapport[_huidig_rapport['Datum'].dt.date == dag_dt]
                    diensten = []
                    for _, rij in dag_df.iterrows():
                        d = str(rij.get('Dienst(en) realisatie', '') or '').strip()
                        naam = str(rij.get('Naam', '') or '').strip()
                        if _re.search(r'\d{1,2}[.:]\d{2}', d) and naam:
                            diensten.append({'naam': naam, 'dienst': d})
                    self.send_json({'diensten': diensten})
                except Exception as e:
                    self.send_json({'diensten': [], 'error': str(e)})
            elif path == '/api/dashboard':
                if _huidig_rapport is None or _huidig_label is None:
                    self.send_json({'geladen': False}); return
                df = _huidig_rapport
                label = _huidig_label
                periode = 'onbekend'
                if 'Datum' in df.columns and df['Datum'].notna().any():
                    periode = f"{df['Datum'].min().strftime('%d-%m-%Y')} – {df['Datum'].max().strftime('%d-%m-%Y')}"
                alarm = []; waarschuwing = []; maand = []
                dagen_met_tekort = 0; totaal_dagen = 0
                try:
                    from datetime import date, timedelta
                    if label not in _rooster_cache:
                        _rooster_cache[label] = analyseer_rooster(df, _medewerkers, _dienst_normen)
                    rc = _rooster_cache[label]
                    vandaag = date.today()
                    over2  = vandaag + timedelta(days=2)
                    over7  = vandaag + timedelta(days=7)
                    over30 = vandaag + timedelta(days=30)
                    for datum_str, dag_info in sorted(rc.items()):
                        totaal_dagen += 1
                        open_groen  = dag_info.get('open_groen', [])
                        open_oranje = dag_info.get('open_oranje', [])
                        alle = open_groen + open_oranje
                        if alle: dagen_met_tekort += 1
                        try: dag_dt = date.fromisoformat(datum_str)
                        except: continue
                        if dag_dt < vandaag: continue
                        if dag_dt <= over2 and open_groen:
                            alarm.append({'datum': datum_str, 'weekdag': dag_info.get('weekdag',''), 'groen': open_groen, 'oranje': open_oranje})
                        elif dag_dt <= over7 and alle:
                            waarschuwing.append({'datum': datum_str, 'weekdag': dag_info.get('weekdag',''), 'groen': open_groen, 'oranje': open_oranje})
                        elif dag_dt <= over30 and alle:
                            maand.append({'datum': datum_str, 'weekdag': dag_info.get('weekdag',''), 'groen': open_groen, 'oranje': open_oranje})
                except Exception as ex:
                    import traceback; print(traceback.format_exc())
                # Vandaag op de vloer — gebruik eerste dag van rapport als vandaag er niet in zit
                import re as _re
                vandaag_diensten = []
                vloer_datum = vandaag
                try:
                    dag_df = df[df['Datum'].dt.date == vandaag]
                    if dag_df.empty:
                        vloer_datum = df['Datum'].min().date()
                        dag_df = df[df['Datum'].dt.date == vloer_datum]
                    for _, rij in dag_df.iterrows():
                        d = str(rij.get('Dienst(en) realisatie', '') or '').strip()
                        naam = str(rij.get('Naam', '') or '').strip()
                        if _re.search(r'\d{1,2}[.:]\d{2}', d) and naam:
                            vandaag_diensten.append({'naam': naam, 'dienst': d})
                except: pass

                # Heatmap komende 14 dagen
                heatmap = []
                try:
                    from datetime import timedelta
                    for i in range(14):
                        dag_dt = vandaag + timedelta(days=i)
                        info = rc.get(dag_dt.isoformat(), {})
                        groen = len(info.get('open_groen', []))
                        oranje = len(info.get('open_oranje', []))
                        heatmap.append({
                            'datum': dag_dt.isoformat(),
                            'weekdag': info.get('weekdag', dag_dt.strftime('%A')),
                            'open_groen': groen,
                            'open_oranje': oranje,
                        })
                except: pass

                # Traject signalen — vrijgekomen plekken voor actieve trajecten
                traject_signalen = []
                try:
                    actieve = laad_actieve_trajecten()
                    if actieve:
                        from maandoverzicht import bouw_omgekeerde_lookup
                        lookup = bouw_omgekeerde_lookup(_medewerkers)
                        traject_signalen = signaleer_vrijgekomen_plekken(df, _medewerkers, lookup, actieve)
                except: pass

                self.send_json({
                    'geladen': True, 'label': label, 'periode': periode,
                    'rijen': len(df), 'alarm': alarm, 'waarschuwing': waarschuwing,
                    'maand': maand, 'dagen_met_tekort': dagen_met_tekort,
                    'totaal_dagen': totaal_dagen,
                    'vandaag_diensten': vandaag_diensten,
                    'vloer_datum': vloer_datum.isoformat(),
                    'heatmap': heatmap,
                    'traject_signalen': traject_signalen,
                })
            elif path == '/api/config':
                self.send_json({
                    'medewerkers_geladen': len(_medewerkers) > 0,
                    'normen_geladen':      len(_dienst_normen) > 0,
                    'aantal_mw':          len(_medewerkers),
                    'aantal_diensten':    len(_dienst_normen),
                })
            elif path == '/api/medewerkers':
                # Geef lijst van medewerkers terug voor weergave
                mw_lijst = []
                for ns, mw in _medewerkers.items():
                    mw_lijst.append({
                        'naam':     mw['naam'],
                        'functie':  mw['functie'],
                        'skills':   mw['skills'],
                        'nacht_ontheffing': mw['nacht_ontheffing'],
                        'is_student': mw['is_student'],
                        'bijzonderheden': mw['bijzonderheden'],
                        'voorkeuren': mw['voorkeuren'],
                    })
                self.send_json({'medewerkers': mw_lijst})
            elif path == '/api/dienstverdeling':
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                df_gebruik = rapporten.get(label) if label else _huidig_rapport
                if df_gebruik is None:
                    self.send_json({'error': 'Geen rapport geladen'}, 400); return
                result = bereken_dienstverdeling(df_gebruik, _uitsluitingen)
                result.pop('df', None)
                self.send_json(result)
            elif path == '/api/senioren':
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                df_gebruik = rapporten.get(label) if label else _huidig_rapport
                if df_gebruik is None:
                    self.send_json({'error': 'Geen rapport geladen'}, 400); return
                self.send_json(bereken_senioren_overzicht(df_gebruik))
            elif path == '/api/koppeling':
                qs = parse_qs(parsed.query)
                mw_ns = qs.get('mw', [None])[0]
                label = qs.get('label', [None])[0]
                fmt   = qs.get('format', ['excel'])[0]
                if not mw_ns:
                    self.send_json({'error': 'Geen medewerker opgegeven'}, 400); return
                import pandas as pd
                # Zoek rapport - ook met strip voor robuustheid
                df_gebruik = None
                if label:
                    label_strip = label.strip()
                    if label_strip in rapporten:
                        df_gebruik = rapporten[label_strip]
                    else:
                        # Probeer gedeeltelijke match
                        for k in rapporten:
                            if label_strip in k or k in label_strip:
                                df_gebruik = rapporten[k]
                                break
                if df_gebruik is None and _jaar_rapport is not None:
                    df_gebruik = _jaar_rapport
                if df_gebruik is None and rapporten:
                    df_gebruik = pd.concat(list(rapporten.values()), ignore_index=True)
                if df_gebruik is None:
                    self.send_json({'error': 'Geen rapport geladen. Laad eerst een rapport via het Rapporten tabblad.'}, 400); return
                data = bouw_koppeling_data(df_gebruik, _medewerkers, mw_ns)
                if fmt == 'json':
                    self.send_json({'data': data, 'totaal': len(data)})
                else:
                    # Zoek naam van kandidaat
                    mw_naam = _medewerkers.get(mw_ns, {}).get('naam', mw_ns)
                    self._exporteer_koppeling(data, mw_naam, fmt)
            elif path == '/api/traject_config':
                # Geef beschikbare medewerkers en trajecten terug
                mw_lijst = [{'ns': ns, 'naam': mw['naam']} 
                            for ns, mw in sorted(_medewerkers.items(), key=lambda x: x[1]['naam'])]
                self.send_json({
                    'medewerkers': mw_lijst,
                    'trajecten':   {k: {'naam': v['naam'], 'omschrijving': v['omschrijving']} 
                                   for k, v in TRAJECTEN.items()},
                    'jaar_rapport_geladen': _jaar_rapport is not None,
                })
            elif path == '/api/actieve_trajecten':
                self.send_json(laad_actieve_trajecten())
            elif path == '/api/traject_kansen':
                qs = parse_qs(parsed.query)
                idx = qs.get('idx', [None])[0]
                actieve = laad_actieve_trajecten()
                if idx is None or _huidig_rapport is None:
                    self.send_json({'error': 'Geen rapport of traject'}); return
                try:
                    actief = actieve[int(idx)]
                    from maandoverzicht import bouw_omgekeerde_lookup
                    lookup = bouw_omgekeerde_lookup(_medewerkers)
                    res = analyseer_traject_kansen(
                        _huidig_rapport, _medewerkers, lookup, actief,
                        al_gedaan_fase1=actief.get('fase1_gedaan', 0),
                        al_gedaan_fase2=actief.get('fase2_gedaan', 0),
                    )
                    self.send_json(res)
                except Exception as e:
                    import traceback; print(traceback.format_exc())
                    self.send_json({'error': str(e)})
            elif path == '/api/traject_analyse':
                qs = parse_qs(parsed.query)
                mw_ns       = qs.get('mw', [None])[0]
                traject_key = qs.get('traject', ['b_spoed'])[0]
                if not mw_ns:
                    self.send_json({'error': 'Geen medewerker opgegeven'}, 400); return
                if _jaar_rapport is None and not rapporten:
                    self.send_json({'error': 'Geen rapport geladen'}, 400); return
                from maandoverzicht import bouw_omgekeerde_lookup
                lookup = bouw_omgekeerde_lookup(_medewerkers)
                # Gebruik jaarrapport als beschikbaar, anders alle geladen rapporten
                import pandas as pd
                if _jaar_rapport is not None:
                    df_gebruik = _jaar_rapport
                elif rapporten:
                    df_gebruik = pd.concat(list(rapporten.values()), ignore_index=True)
                else:
                    self.send_json({'error': 'Geen data beschikbaar'}, 400); return
                res = analyseer_traject(df_gebruik, _medewerkers, mw_ns, traject_key, lookup)
                sam = samenvatting(res, traject_key)
                self.send_json(sam)
            elif path == '/api/traject_schema':
                qs = parse_qs(parsed.query)
                mw_ns       = qs.get('mw', [None])[0]
                traject_key = qs.get('traject', ['b_spoed'])[0]
                label       = qs.get('label', [None])[0]
                if not mw_ns or not label or label not in rapporten:
                    self.send_json({'error': 'Selecteer een medewerker en een rapport.'}, 400); return
                from maandoverzicht import bouw_omgekeerde_lookup
                import pandas as pd
                lookup = bouw_omgekeerde_lookup(_medewerkers)
                # Historische data: jaarrapport OF alle andere rapporten behalve het gekozen
                if _jaar_rapport is not None:
                    df_hist = _jaar_rapport
                elif len(rapporten) > 1:
                    df_hist = pd.concat([v for k,v in rapporten.items() if k != label], ignore_index=True)
                else:
                    df_hist = rapporten[label]  # gebruik hetzelfde rapport als er maar 1 is
                df_schema = rapporten[label]
                res    = analyseer_traject(df_hist, _medewerkers, mw_ns, traject_key, lookup) if not df_hist.empty else []
                sam    = samenvatting(res, traject_key)
                schema = genereer_inwerkschema(df_schema, _medewerkers, mw_ns, traject_key, lookup, sam)
                self.send_json({'samenvatting': sam, 'schema': schema})
            elif path == '/api/maandoverzicht_mw':
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                df_gebruik = rapporten.get(label) if label else _huidig_rapport
                if df_gebruik is None:
                    self.send_json({'error': 'Geen rapport geladen'}, 400); return
                data = bouw_maandoverzicht(df_gebruik, _medewerkers)
                self.send_json(data)
            elif path == '/api/werkplek':
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                df_gebruik = rapporten.get(label) if label else _huidig_rapport
                if df_gebruik is None:
                    self.send_json({'error': 'Geen rapport geladen'}, 400); return
                data = bouw_maand_werkplek(df_gebruik)
                self.send_json(data)
            elif path == '/api/maandmatrix':
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                df_gebruik = rapporten.get(label) if label else _huidig_rapport
                lbl_key = label or _huidig_label or 'huidig'
                if df_gebruik is None:
                    self.send_json({'error': 'Geen rapport geladen'}, 400); return
                if lbl_key not in _rooster_cache:
                    _rooster_cache[lbl_key] = analyseer_rooster(df_gebruik, _medewerkers, _dienst_normen)
                matrix = bouw_matrix(_rooster_cache[lbl_key], _dienst_normen)
                self.send_json(matrix)
            elif path == '/api/roostercheck':
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                if not label or label not in rapporten:
                    self.send_json({'error': 'Rapport niet gevonden'}, 400); return
                if not _medewerkers:
                    self.send_json({'error': 'Medewerkers nog niet geladen. Upload eerst het medewerkers-bestand.'}, 400); return
                if not _dienst_normen:
                    self.send_json({'error': 'Dienst normen nog niet geladen. Upload eerst het normen-bestand.'}, 400); return
                # Cache
                if label not in _rooster_cache:
                    _rooster_cache[label] = analyseer_rooster(rapporten[label], _medewerkers, _dienst_normen)
                self.send_json(_rooster_cache[label])
            elif path == '/api/vergelijk':
                qs = parse_qs(parsed.query)
                l1 = qs.get('label1', [None])[0]
                l2 = qs.get('label2', [None])[0]
                if not l1 or not l2 or l1 not in rapporten or l2 not in rapporten:
                    self.send_json({'error': 'Rapporten niet gevonden'}, 400); return
                result = vergelijk_rapporten(rapporten[l1], l1, rapporten[l2], l2, _uitsluitingen)
                result['verdeling1'].pop('df', None)
                result['verdeling2'].pop('df', None)
                self.send_json(result)
            elif path == '/api/exporteer_tekorten':
                qs = parse_qs(parsed.query)
                label  = qs.get('label',  [None])[0]
                labels = qs.get('labels', [None])[0]
                fmt    = qs.get('format', ['excel'])[0]
                if labels:
                    labels_list = [l for l in labels.split(',') if l in rapporten]
                    if not labels_list:
                        self.send_json({'error': 'Geen geldige rapporten'}, 400); return
                    for lbl in labels_list:
                        if lbl not in _rooster_cache:
                            _rooster_cache[lbl] = analyseer_rooster(rapporten[lbl], _medewerkers, _dienst_normen)
                    self._exporteer_tekorten(labels_list[0], fmt, labels=labels_list)
                elif label:
                    if label not in rapporten:
                        self.send_json({'error': 'Rapport niet gevonden'}, 400); return
                    if not _dienst_normen:
                        self.send_json({'error': 'Normen niet geladen'}, 400); return
                    if label not in _rooster_cache:
                        _rooster_cache[label] = analyseer_rooster(rapporten[label], _medewerkers, _dienst_normen)
                    self._exporteer_tekorten(label, fmt)
                else:
                    self.send_json({'error': 'Geen label opgegeven'}, 400); return
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                fmt   = qs.get('format', ['excel'])[0]
                if not label or label not in rapporten:
                    self.send_json({'error': 'Rapport niet gevonden'}, 400); return
                if not _dienst_normen:
                    self.send_json({'error': 'Normen niet geladen'}, 400); return
                if label not in _rooster_cache:
                    _rooster_cache[label] = analyseer_rooster(rapporten[label], _medewerkers, _dienst_normen)
                self._exporteer_tekorten(label, fmt)
            elif path == '/api/exporteer':
                qs = parse_qs(parsed.query)
                label = qs.get('label', [None])[0]
                fmt   = qs.get('format', ['excel'])[0]
                if not label or label not in rapporten:
                    self.send_json({'error': 'Rapport niet gevonden'}, 400); return
                self._exporteer(label, fmt)
            elif path == '/stop':
                self.send_html('<h2>App afgesloten.</h2>')
                threading.Thread(target=lambda: os._exit(0), daemon=True).start()
            else:
                self.send_json({'error': 'Niet gevonden'}, 404)
        except Exception as e:
            self.send_json({'error': str(e), 'trace': traceback.format_exc()}, 500)

    def do_POST(self):
        parsed = urlparse(self.path)
        path   = parsed.path
        try:
            if path == '/api/upload':
                self._upload()
            elif path == '/api/upload_jaar':
                self._upload_jaar()
            elif path == '/api/upload_config':
                self._upload_config()
            elif path == '/api/verwijder':
                self._verwijder()
            elif path == '/api/actief_traject_toevoegen':
                length = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(length))
                actieve = laad_actieve_trajecten()
                # Voorkom dubbelen
                al = any(a['naam_schoon'] == body['naam_schoon'] and a['traject'] == body['traject']
                         for a in actieve)
                if not al:
                    actieve.append({
                        'naam':        body['naam'],
                        'naam_schoon': body['naam_schoon'],
                        'traject':     body['traject'],
                        'fase1_gedaan': 0,
                        'fase2_gedaan': 0,
                        'gestart':     body.get('gestart', ''),
                    })
                    sla_actieve_trajecten_op(actieve)
                self.send_json({'ok': True, 'trajecten': actieve})
            elif path == '/api/actief_traject_verwijderen':
                length = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(length))
                actieve = laad_actieve_trajecten()
                actieve = [a for i, a in enumerate(actieve) if i != int(body['idx'])]
                sla_actieve_trajecten_op(actieve)
                self.send_json({'ok': True, 'trajecten': actieve})
            elif path == '/api/actief_traject_update':
                length = int(self.headers.get('Content-Length', 0))
                body = json.loads(self.rfile.read(length))
                actieve = laad_actieve_trajecten()
                idx = int(body['idx'])
                if 0 <= idx < len(actieve):
                    actieve[idx]['fase1_gedaan'] = int(body.get('fase1_gedaan', 0))
                    actieve[idx]['fase2_gedaan'] = int(body.get('fase2_gedaan', 0))
                    sla_actieve_trajecten_op(actieve)
                self.send_json({'ok': True, 'trajecten': actieve})
            else:
                self.send_json({'error': 'Niet gevonden'}, 404)
        except Exception as e:
            self.send_json({'error': str(e), 'trace': traceback.format_exc()}, 500)

    def _upload_jaar(self):
        """Upload en sla jaarrapport permanent op."""
        global _jaar_rapport
        import tempfile, re
        ctype  = self.headers.get('Content-Type', '')
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length)
        boundary_match = re.search(r'boundary=([^\s;]+)', ctype)
        if not boundary_match:
            self.send_json({'error': 'Geen boundary'}, 400); return
        boundary = boundary_match.group(1).encode()
        bestand_bytes = None
        for part in body.split(b'--' + boundary):
            if b'Content-Disposition' not in part: continue
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1: continue
            header = part[:header_end].decode('utf-8', errors='replace')
            inhoud = part[header_end + 4:]
            if inhoud.endswith(b'\r\n'): inhoud = inhoud[:-2]
            import re as _re
            if _re.search(r'filename=', header):
                bestand_bytes = inhoud
        if not bestand_bytes:
            self.send_json({'error': 'Geen bestand'}, 400); return
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(bestand_bytes); tmp_pad = tmp.name
        try:
            from analyse import laad_rapport
            df = laad_rapport(tmp_pad)
            _jaar_rapport = df
            # Sla ook op als xlsx voor hergebruik na herstart
            import shutil, os
            dst = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data_jaarrapport.xlsx')
            shutil.copy(tmp_pad, dst)
            periode = 'onbekend'
            if 'Datum' in df.columns and df['Datum'].notna().any():
                periode = f"{df['Datum'].min().strftime('%d-%m-%Y')} – {df['Datum'].max().strftime('%d-%m-%Y')}"
            self.send_json({'ok': True, 'rijen': len(df), 'periode': periode})
        except Exception as e:
            self.send_json({'error': str(e)}, 500)
        finally:
            import os; os.unlink(tmp_pad)

    def _upload(self):
        import tempfile, re
        ctype  = self.headers.get('Content-Type', '')
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length)

        boundary_match = re.search(r'boundary=([^\s;]+)', ctype)
        if not boundary_match:
            self.send_json({'error': 'Geen boundary'}, 400); return
        boundary = boundary_match.group(1).encode()

        bestandsnaam = None; bestand_bytes = None; label_custom = None
        for part in body.split(b'--' + boundary):
            if b'Content-Disposition' not in part: continue
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1: continue
            header = part[:header_end].decode('utf-8', errors='replace')
            inhoud = part[header_end + 4:]
            if inhoud.endswith(b'\r\n'): inhoud = inhoud[:-2]
            fn = re.search(r'filename="([^"]+)"', header)
            nm = re.search(r'name="([^"]+)"', header)
            if fn:
                bestandsnaam = fn.group(1); bestand_bytes = inhoud
            elif nm and nm.group(1) == 'label':
                label_custom = inhoud.decode('utf-8').strip()

        if not bestand_bytes:
            self.send_json({'error': 'Geen bestand'}, 400); return

        suffix = '.xlsx' if bestandsnaam and 'xlsx' in bestandsnaam else '.xls'
        with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(bestand_bytes); tmp_pad = tmp.name

        try:
            df = laad_rapport(tmp_pad)
            label = label_custom or bestandsnaam or 'rapport'
            base_label = label; i = 1
            while label in rapporten:
                label = f"{base_label} ({i})"; i += 1
            rapporten[label] = df
            # Invalideer cache
            _rooster_cache.pop(label, None)
            periode = 'onbekend'
            if 'Datum' in df.columns and df['Datum'].notna().any():
                periode = f"{df['Datum'].min().strftime('%d-%m-%Y')} – {df['Datum'].max().strftime('%d-%m-%Y')}"
            # Zet als huidig actief rapport
            global _huidig_rapport, _huidig_label
            _huidig_rapport = df
            _huidig_label   = label
            # Start roostercheck alvast in de achtergrond zodat dashboard snel laadt
            def _warmup():
                try:
                    if label not in _rooster_cache:
                        _rooster_cache[label] = analyseer_rooster(df, _medewerkers, _dienst_normen)
                except: pass
            threading.Thread(target=_warmup, daemon=True).start()
            self.send_json({'ok': True, 'label': label, 'rijen': len(df), 'periode': periode})
        finally:
            os.unlink(tmp_pad)

    def _upload_config(self):
        """Upload medewerkers of normen bestand."""
        global _medewerkers, _dienst_normen, _rooster_cache
        import tempfile, re
        ctype  = self.headers.get('Content-Type', '')
        length = int(self.headers.get('Content-Length', 0))
        body   = self.rfile.read(length)

        boundary_match = re.search(r'boundary=([^\s;]+)', ctype)
        if not boundary_match:
            self.send_json({'error': 'Geen boundary'}, 400); return
        boundary = boundary_match.group(1).encode()

        bestandsnaam = None; bestand_bytes = None; type_config = None
        for part in body.split(b'--' + boundary):
            if b'Content-Disposition' not in part: continue
            header_end = part.find(b'\r\n\r\n')
            if header_end == -1: continue
            header = part[:header_end].decode('utf-8', errors='replace')
            inhoud = part[header_end + 4:]
            if inhoud.endswith(b'\r\n'): inhoud = inhoud[:-2]
            fn = re.search(r'filename="([^"]+)"', header)
            nm = re.search(r'name="([^"]+)"', header)
            if fn:
                bestandsnaam = fn.group(1); bestand_bytes = inhoud
            elif nm and nm.group(1) == 'type':
                type_config = inhoud.decode('utf-8').strip()

        if not bestand_bytes:
            self.send_json({'error': 'Geen bestand'}, 400); return

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(bestand_bytes); tmp_pad = tmp.name

        try:
            if type_config == 'medewerkers':
                _medewerkers = laad_medewerkers(tmp_pad)
                _rooster_cache.clear()
                self.send_json({'ok': True, 'type': 'medewerkers', 'aantal': len(_medewerkers)})
            elif type_config == 'normen':
                _dienst_normen = laad_dienst_normen(tmp_pad)
                _rooster_cache.clear()
                self.send_json({'ok': True, 'type': 'normen', 'aantal': len(_dienst_normen)})
            else:
                self.send_json({'error': f'Onbekend type: {type_config}'}, 400)
        except Exception as e:
            self.send_json({'error': str(e)}, 500)
        finally:
            os.unlink(tmp_pad)

    def _verwijder(self):
        length = int(self.headers.get('Content-Length', 0))
        data = json.loads(self.rfile.read(length))
        label = data.get('label')
        if label in rapporten:
            del rapporten[label]
            _rooster_cache.pop(label, None)
            self.send_json({'ok': True})
        else:
            self.send_json({'error': 'Niet gevonden'}, 404)

    def _exporteer_koppeling(self, data, mw_naam, fmt):
        import io
        if not data:
            self.send_json({'error': 'Geen diensten gevonden'}); return

        if fmt == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            buf = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Koppelingsoverzicht'

            # Titel
            ws.merge_cells('A1:C1')
            ws['A1'] = f'Koppelingsoverzicht — {mw_naam}'
            ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
            ws['A1'].fill = PatternFill('solid', fgColor='1F4E79')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 26

            # Headers
            for col, (tekst, breedte) in enumerate([('Datum',32),('Dienst',22),('Collegas op zelfde dienst',40)], 1):
                cel = ws.cell(2, col, tekst)
                cel.font = Font(bold=True, color='FFFFFF', size=11)
                cel.fill = PatternFill('solid', fgColor='2E75B6')
                cel.alignment = Alignment(horizontal='center' if col > 1 else 'left')
                ws.column_dimensions[chr(64+col)].width = breedte
            ws.row_dimensions[2].height = 20

            thin = Side(style='thin', color='BCC4CC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for i, r in enumerate(data):
                rij = i + 3
                ws.cell(rij, 1, r['datum_nl']).border = border
                ws.cell(rij, 2, r['dienst']).border = border
                ws.cell(rij, 3, ', '.join(r['collegas']) if r['collegas'] else '—').border = border
                if i % 2 == 0:
                    for col in range(1, 4):
                        ws.cell(rij, col).fill = PatternFill('solid', fgColor='F5F7FA')
                ws.row_dimensions[rij].height = 18

            body = io.BytesIO()
            wb.save(body)
            body = body.getvalue()
            fname = f"Koppeling_{mw_naam.replace(' ','_')}.xlsx"
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)

        elif fmt == 'pdf':
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4,
                leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            story = []
            titel_style = ParagraphStyle('titel', parent=styles['Title'],
                fontSize=14, textColor=colors.HexColor('#1F4E79'), spaceAfter=4)
            story.append(Paragraph(f'Koppelingsoverzicht — {mw_naam}', titel_style))
            story.append(Paragraph(f'{len(data)} diensten gevonden.',
                ParagraphStyle('sub', parent=styles['Normal'], fontSize=10,
                    textColor=colors.HexColor('#7A8899'), spaceAfter=12)))
            tabel_data = [['Datum', 'Dienst', "Collega's op zelfde dienst"]]
            stijlen = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR', (0,0), (-1,0), colors.white),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,-1), 9),
                ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#BCC4CC')),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING', (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#F5F7FA'), colors.white]),
            ]
            for r in data:
                tabel_data.append([r['datum_nl'], r['dienst'], ', '.join(r['collegas']) if r['collegas'] else '—'])
            t = Table(tabel_data, colWidths=[6*cm, 4*cm, 7.5*cm], repeatRows=1)
            t.setStyle(TableStyle(stijlen))
            story.append(t)
            doc.build(story)
            body = buf.getvalue()
            fname = f"Koppeling_{mw_naam.replace(' ','_')}.pdf"
            self.send_response(200)
            self.send_header('Content-Type', 'application/pdf')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)

    def _exporteer_tekorten(self, label, fmt, labels=None):
        import io
        # Ondersteuning voor meerdere maanden
        if labels:
            alle_rijen = []
            for lbl in labels:
                if lbl in rapporten:
                    if lbl not in _rooster_cache:
                        _rooster_cache[lbl] = analyseer_rooster(rapporten[lbl], _medewerkers, _dienst_normen)
                    alle_rijen.extend(_exporteer_tekorten_data(_rooster_cache[lbl]))
            rijen = sorted(alle_rijen, key=lambda r: (r['datum_str'], r['prio']))
        else:
            rijen = _exporteer_tekorten_data(_rooster_cache[label])

        if not rijen:
            self.send_json({'error': 'Geen tekorten gevonden'}); return

        # Kleur per prioriteit
        KLEUREN = {
            0: {'bg': 'FFFFFF', 'fg': '1A2530', 'bold': False},  # normaal
            1: {'bg': 'E6F1FB', 'fg': '0C447C', 'bold': False},  # nacht
            2: {'bg': 'FAEEDA', 'fg': '633806', 'bold': True},   # A dienst
            3: {'bg': 'A32D2D', 'fg': 'F7C1C1', 'bold': True},   # A nacht
        }
        DAGDEEL_NL = {'dag': 'dag', 'avond': 'avond', 'nacht': 'nacht'}

        if fmt == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            buf = io.BytesIO()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Tekorten'

            # Titel
            lbl_tekst = ', '.join(labels) if labels else label
            ws.merge_cells('A1:D1')
            ws['A1'] = f'Openstaande diensten — {lbl_tekst}'
            ws['A1'].font = Font(bold=True, size=13, color='FFFFFF')
            ws['A1'].fill = PatternFill('solid', fgColor='1F4E79')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 26

            # Legenda rijen
            legenda = [
                ('Nachtdienst', 'E6F1FB', '0C447C'),
                ('A dienst', 'FAEEDA', '633806'),
                ('A nachtdienst (hoogste prioriteit)', 'A32D2D', 'F7C1C1'),
            ]
            for i, (tekst, bg, fg) in enumerate(legenda):
                r = i + 2
                ws.merge_cells(f'A{r}:D{r}')
                ws[f'A{r}'] = tekst
                ws[f'A{r}'].fill = PatternFill('solid', fgColor=bg)
                ws[f'A{r}'].font = Font(color=fg, size=10, italic=True)
                ws[f'A{r}'].alignment = Alignment(horizontal='left', indent=1)
                ws.row_dimensions[r].height = 16

            # Headers
            header_rij = 5
            for col, (tekst, breedte) in enumerate([('Datum',34),('Dagdeel',12),('Dienst',20),('Naam',28)], 1):
                cel = ws.cell(header_rij, col, tekst)
                cel.font = Font(bold=True, color='FFFFFF', size=11)
                cel.fill = PatternFill('solid', fgColor='2E75B6')
                cel.alignment = Alignment(horizontal='center' if col > 1 else 'left')
                ws.column_dimensions[chr(64+col)].width = breedte
            ws.row_dimensions[header_rij].height = 20

            thin = Side(style='thin', color='BCC4CC')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for i, r in enumerate(rijen):
                kleur = KLEUREN[r['prio']]
                rij_nr = i + header_rij + 1
                waarden = [r['datum_nl'], DAGDEEL_NL[r['dagdeel']], r['dienst'], r['wie']]
                for col, val in enumerate(waarden, 1):
                    cel = ws.cell(rij_nr, col, val)
                    cel.fill = PatternFill('solid', fgColor=kleur['bg'])
                    cel.font = Font(color=kleur['fg'], bold=kleur['bold'] and col in (2,3))
                    cel.border = border
                    cel.alignment = Alignment(vertical='center',
                        horizontal='center' if col == 2 else 'left')
                ws.row_dimensions[rij_nr].height = 18

            body = io.BytesIO()
            wb.save(body)
            body = body.getvalue()
            fname = f"Tekorten_{lbl_tekst.replace(' ','_')[:40]}.xlsx"
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)

        elif fmt == 'pdf':
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm

            buf = io.BytesIO()
            lbl_tekst = ', '.join(labels) if labels else label
            doc = SimpleDocTemplate(buf, pagesize=A4,
                leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=2*cm, bottomMargin=2*cm)
            styles = getSampleStyleSheet()
            story = []

            titel_style = ParagraphStyle('titel', parent=styles['Title'],
                fontSize=14, textColor=colors.HexColor('#1F4E79'), spaceAfter=4)
            story.append(Paragraph(f'Openstaande diensten — {lbl_tekst}', titel_style))
            story.append(Paragraph('Schrijf je naam als je een dienst kunt overnemen.',
                ParagraphStyle('sub', parent=styles['Normal'], fontSize=10,
                    textColor=colors.HexColor('#7A8899'), spaceAfter=12)))

            tabel_data = [['Datum', 'Dagdeel', 'Dienst', 'Naam']]
            stijlen = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')),
                ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
                ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',   (0,0), (-1,0), 9),
                ('ALIGN',      (0,0), (-1,0), 'CENTER'),
                ('BOTTOMPADDING', (0,0), (-1,0), 7),
                ('TOPPADDING',    (0,0), (-1,0), 7),
                ('FONTSIZE',   (0,1), (-1,-1), 9),
                ('GRID',       (0,0), (-1,-1), 0.5, colors.HexColor('#BCC4CC')),
                ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
                ('TOPPADDING', (0,1), (-1,-1), 5),
                ('BOTTOMPADDING', (0,1), (-1,-1), 5),
                ('ALIGN',      (1,1), (1,-1), 'CENTER'),
                ('ALIGN',      (2,1), (2,-1), 'CENTER'),
            ]

            PRIO_KLEUREN_PDF = {
                0: (None, colors.HexColor('#1A2530'), False),
                1: (colors.HexColor('#E6F1FB'), colors.HexColor('#0C447C'), False),
                2: (colors.HexColor('#FAEEDA'), colors.HexColor('#633806'), True),
                3: (colors.HexColor('#A32D2D'), colors.HexColor('#F7C1C1'), True),
            }

            for i, r in enumerate(rijen):
                tabel_data.append([r['datum_nl'], DAGDEEL_NL[r['dagdeel']], r['dienst'], ''])
                rij_nr = i + 1
                bg, fg, bold = PRIO_KLEUREN_PDF[r['prio']]
                if bg:
                    stijlen.append(('BACKGROUND', (0,rij_nr), (-1,rij_nr), bg))
                stijlen.append(('TEXTCOLOR', (0,rij_nr), (-1,rij_nr), fg))
                if bold:
                    stijlen.append(('FONTNAME', (0,rij_nr), (-1,rij_nr), 'Helvetica-Bold'))

            col_breedte = [6.5*cm, 2.2*cm, 4.5*cm, 4.3*cm]
            t = Table(tabel_data, colWidths=col_breedte, repeatRows=1)
            t.setStyle(TableStyle(stijlen))
            story.append(t)
            doc.build(story)

            body = buf.getvalue()
            fname = f"Tekorten_{lbl_tekst.replace(' ','_')[:40]}.pdf"
            self.send_response(200)
            self.send_header('Content-Type', 'application/pdf')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)

    def _exporteer(self, label, fmt):
        import io
        df = rapporten[label]
        result = bereken_dienstverdeling(df)
        overzicht = bereken_senioren_overzicht(df)

        if fmt == 'excel':
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine='openpyxl') as writer:
                rows = [
                    ['Categorie', 'Aantal', '%'],
                    ['Coretaken', result['coretaken']['aantal'], result['coretaken']['pct']],
                    ['Overige diensten', result['overige']['aantal'], result['overige']['pct']],
                    ['x-Diensten', result['xdiensten']['aantal'], result['xdiensten']['pct']],
                    ['TOTAAL', result['totaal'], 100],
                    ['Overwerk (buiten totaal)', result['overwerk']['aantal'], ''],
                ]
                pd.DataFrame(rows[1:], columns=rows[0]).to_excel(writer, sheet_name='Dienstverdeling', index=False)
                senior_rows = []
                for datum_str in sorted(overzicht.keys()):
                    for item in overzicht[datum_str]:
                        senior_rows.append({'Datum': datum_str, 'Naam': item['naam'], 'Dienst': item['dienst']})
                if senior_rows:
                    pd.DataFrame(senior_rows).to_excel(writer, sheet_name='Senioren per dag', index=False)
                # Roostercheck exporteren als ook beschikbaar
                if label in _rooster_cache:
                    check_rows = []
                    for datum_str, dag in sorted(_rooster_cache[label].items()):
                        for open_d in dag['open_groen']:
                            check_rows.append({'Datum': datum_str, 'Weekdag': dag['weekdag'], 'Status': 'OPEN ESSENTIEEL', 'Dienst': open_d})
                        for open_d in dag['open_oranje']:
                            check_rows.append({'Datum': datum_str, 'Weekdag': dag['weekdag'], 'Status': 'Open wenselijk', 'Dienst': open_d})
                    if check_rows:
                        pd.DataFrame(check_rows).to_excel(writer, sheet_name='Roostercheck', index=False)
            body = buf.getvalue()
            fname = f"MKA_export_{label.replace(' ','_')}.xlsx"
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)
        elif fmt == 'pdf':
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            buf = io.BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4)
            styles = getSampleStyleSheet()
            story = [Paragraph(f'MKA Analyse — {label}', styles['Title']), Spacer(1,12),
                     Paragraph('Dienstverdeling', styles['Heading2'])]
            tabel_data = [
                ['Categorie','Aantal','%'],
                ['Coretaken', str(result['coretaken']['aantal']), f"{result['coretaken']['pct']}%"],
                ['Overige', str(result['overige']['aantal']), f"{result['overige']['pct']}%"],
                ['x-Diensten', str(result['xdiensten']['aantal']), f"{result['xdiensten']['pct']}%"],
                ['TOTAAL', str(result['totaal']), '100%'],
            ]
            t = Table(tabel_data, colWidths=[200,80,80])
            t.setStyle(TableStyle([
                ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1F4E79')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('GRID',(0,0),(-1,-1),0.5,colors.grey),
            ]))
            story.append(t)
            doc.build(story)
            body = buf.getvalue()
            fname = f"MKA_export_{label.replace(' ','_')}.pdf"
            self.send_response(200)
            self.send_header('Content-Type', 'application/pdf')
            self.send_header('Content-Disposition', f'attachment; filename="{fname}"')
            self.send_header('Content-Length', len(body))
            self.end_headers()
            self.wfile.write(body)


def _exporteer_tekorten_data(rooster_res):
    """Bouw lijst van tekorten: datum, dienst, lege Wie kolom."""
    rijen = []
    _DAGNAMEN = ['maandag','dinsdag','woensdag','donderdag','vrijdag','zaterdag','zondag']
    _MAANDEN  = ['','januari','februari','maart','april','mei','juni',
                 'juli','augustus','september','oktober','november','december']
    for datum_str in sorted(rooster_res.keys()):
        dag = rooster_res[datum_str]
        from datetime import datetime
        dt = datetime.strptime(datum_str, '%Y-%m-%d')
        datum_nl = f"{_DAGNAMEN[dt.weekday()]} {dt.day} {_MAANDEN[dt.month]} {dt.year}"
        for dienst in dag.get('open_groen', []):
            rijen.append({'datum_nl': datum_nl, 'datum_str': datum_str, 'dienst': dienst, 'wie': ''})
        for dienst in dag.get('open_oranje', []):
            rijen.append({'datum_nl': datum_nl, 'datum_str': datum_str, 'dienst': dienst, 'wie': ''})
    return rijen

def _dagdeel(dienst):
    """Geef dagdeel terug op basis van begintijd."""
    import re as _re
    m = _re.search(r'(\d{1,2})[:.](\d{2})', dienst or '')
    if not m:
        return 'dag'
    uur = int(m.group(1))
    if uur >= 22 or uur < 6:
        return 'nacht'
    if uur >= 14:
        return 'avond'
    return 'dag'

def _prioriteit(dienst):
    """Geef prioriteit terug: 0=normaal, 1=nacht, 2=A dienst, 3=A nacht."""
    is_a     = bool(__import__('re').search(r'\bA\b', dienst or ''))
    is_nacht = _dagdeel(dienst) == 'nacht'
    if is_a and is_nacht: return 3
    if is_a:              return 2
    if is_nacht:          return 1
    return 0

def _exporteer_tekorten_data(rooster_res):
    """Bouw lijst van tekorten gesorteerd op datum, dan prioriteit."""
    rijen = []
    _DAGNAMEN = ['maandag','dinsdag','woensdag','donderdag','vrijdag','zaterdag','zondag']
    _MAANDEN  = ['','januari','februari','maart','april','mei','juni',
                 'juli','augustus','september','oktober','november','december']
    from datetime import datetime
    for datum_str in sorted(rooster_res.keys()):
        dag = rooster_res[datum_str]
        dt = datetime.strptime(datum_str, '%Y-%m-%d')
        datum_nl = f"{_DAGNAMEN[dt.weekday()]} {dt.day} {_MAANDEN[dt.month]} {dt.year}"
        for dienst in dag.get('open_groen', []) + dag.get('open_oranje', []):
            rijen.append({
                'datum_nl':  datum_nl,
                'datum_str': datum_str,
                'dienst':    dienst,
                'dagdeel':   _dagdeel(dienst),
                'prio':      _prioriteit(dienst),
                'wie':       '',
            })
    return rijen


def bouw_koppeling_data(df, mw_db, kandidaat_ns):
    """Bouw koppelingsoverzicht: per dag de dienst en collega's op dezelfde dienst."""
    import pandas as pd, re
    from rooster import normaliseer_dienst, extraheer_dienst
    from maandoverzicht import zoek_mw, bouw_omgekeerde_lookup
    from analyse import naam_schoon

    lookup = bouw_omgekeerde_lookup(mw_db)
    _DAGNAMEN = ['maandag','dinsdag','woensdag','donderdag','vrijdag','zaterdag','zondag']
    _MAANDEN  = ['','januari','februari','maart','april','mei','juni',
                 'juli','augustus','september','oktober','november','december']

    resultaten = []
    for datum, dag_df in df.groupby('Datum'):
        if pd.isna(datum): continue

        # Zoek dienst van kandidaat
        kandidaat_dienst = None
        for _, r in dag_df.iterrows():
            naam = str(r.get('Naam','') or '').strip()
            mw = zoek_mw(naam, lookup)
            if not mw or mw['naam_schoon'] != kandidaat_ns: continue
            d_raw = extraheer_dienst(str(r.get('Dienst(en) realisatie','') or '').strip())
            dn = normaliseer_dienst(d_raw) or d_raw
            inzet = str(r.get('Inzet','') or '')
            if dn and dn != '-' and re.search(r'\d{1,2}:\d{2}', dn):
                if not (inzet == 'Inzet 2' and not re.search(r'\d{1,2}[.:]\d{2}', d_raw)):
                    kandidaat_dienst = dn
                    break

        if not kandidaat_dienst: continue

        # Zoek collega's op zelfde dienst
        collegas = []
        for _, r in dag_df.iterrows():
            naam = str(r.get('Naam','') or '').strip()
            mw = zoek_mw(naam, lookup)
            if not mw or mw['naam_schoon'] == kandidaat_ns: continue
            d_raw = extraheer_dienst(str(r.get('Dienst(en) realisatie','') or '').strip())
            dn = normaliseer_dienst(d_raw) or d_raw
            dn_clean = dn.lstrip('x').strip()
            inzet = str(r.get('Inzet','') or '')
            if not dn or dn == '-': continue
            if not (inzet == 'Inzet 2' and not re.search(r'\d{1,2}[.:]\d{2}', d_raw)):
                if dn_clean == kandidaat_dienst or dn == kandidaat_dienst:
                    collegas.append(mw['naam'])

        dt = pd.Timestamp(datum)
        datum_nl = f"{_DAGNAMEN[dt.weekday()]} {dt.day} {_MAANDEN[dt.month]} {dt.year}"

        resultaten.append({
            'datum_str': dt.strftime('%Y-%m-%d'),
            'datum_nl':  datum_nl,
            'dienst':    kandidaat_dienst,
            'collegas':  collegas,
        })

    return resultaten

def bouw_matrix(rooster_res, dienst_normen):
    """Bouw matrix data: diensten x dagen, met status per cel."""
    datums = sorted(rooster_res.keys())
    if not datums:
        return {'datums': [], 'diensten': [], 'cellen': {}}

    # Bepaal alle normdiensten gesorteerd op prioriteit (groen eerst)
    # Gebruik eerste dag als referentie maar combineer alle dagen
    prioriteit = {}  # dienst -> beste status over de week
    for dienst, dagen in dienst_normen.items():
        statussen = [d.get('status', 'rood') for d in dagen.values()]
        if 'groen' in statussen:
            prioriteit[dienst] = 0
        elif 'oranje' in statussen:
            prioriteit[dienst] = 1
        else:
            prioriteit[dienst] = 2

    diensten_gesorteerd = sorted(prioriteit.keys(), key=lambda d: (prioriteit[d], d))

    # Bouw cellen: datum -> dienst -> {status_norm, gevuld, namen, is_open, is_rood_gevuld}
    cellen = {}
    for datum_str in datums:
        dag = rooster_res[datum_str]
        weekdag = dag['weekdag']
        cellen[datum_str] = {}

        for dienst in diensten_gesorteerd:
            norm = dienst_normen[dienst].get(weekdag, {})
            status_norm = norm.get('status', 'rood')

            # Wie staat er op deze dienst?
            ingepland = dag['ingepland'].get(dienst, [])
            namen = [p['naam'].split()[-1] for p in ingepland if not p.get('is_x') and not p.get('is_overwerk')]

            is_open = dienst in dag['open_groen'] or dienst in dag['open_oranje']
            is_rood_gevuld = len(namen) > 0 and status_norm == 'rood'
            is_essentiel_open = dienst in dag['open_groen']
            is_wenselijk_open = dienst in dag['open_oranje']

            cellen[datum_str][dienst] = {
                'status_norm':      status_norm,
                'namen':            namen,
                'gevuld':           len(namen) > 0,
                'is_essentiel_open': is_essentiel_open,
                'is_wenselijk_open': is_wenselijk_open,
                'is_rood_gevuld':    is_rood_gevuld,
            }

    # Dagsamenvattingen voor hittekaartrij
    dag_info = {}
    for datum_str in datums:
        dag = rooster_res[datum_str]
        dag_info[datum_str] = {
            'weekdag':    dag['weekdag'],
            'tekort':     dag['tekort'],
            'open_oranje': len(dag['open_oranje']),
            'rood_gevuld': len(dag['gevuld_rood']),
        }

    return {
        'datums':   datums,
        'diensten': diensten_gesorteerd,
        'prioriteit': prioriteit,
        'cellen':   cellen,
        'dag_info': dag_info,
    }

def start_server(poort=5757):
    global _medewerkers, _dienst_normen

    # Auto-laad vaste bestanden
    base = os.path.dirname(os.path.abspath(__file__))
    mw_pad      = os.path.join(base, 'data_medewerkers.xlsx')
    normen_pad  = os.path.join(base, 'data_normen.xlsx')
    base = os.path.dirname(os.path.abspath(__file__))

    uitsl_pad  = os.path.join(base, 'data_uitsluitingen.xlsx')
    mw_pad     = os.path.join(base, 'data_medewerkers.xlsx')
    normen_pad = os.path.join(base, 'data_normen.xlsx')

    jaar_pad = os.path.join(base, 'data_jaarrapport.xlsx')
    if os.path.exists(jaar_pad):
        try:
            from analyse import laad_rapport as _lr
            _jaar_rapport = _lr(jaar_pad)
            print(f"  ✓ Jaarrapport geladen ({len(_jaar_rapport)} rijen, {_jaar_rapport['Datum'].min().strftime('%d-%m-%Y')} – {_jaar_rapport['Datum'].max().strftime('%d-%m-%Y')})")
        except Exception as e:
            print(f"  ✗ Jaarrapport laden mislukt: {e}")

    if os.path.exists(uitsl_pad):
        try:
            _uitsluitingen = laad_uitsluitingen(uitsl_pad)
            print(f"  ✓ {len(_uitsluitingen)} uitsluitingen geladen")
        except Exception as e:
            print(f"  ✗ Uitsluitingen laden mislukt: {e}")

    if os.path.exists(mw_pad):
        try:
            _medewerkers = laad_medewerkers(mw_pad)
            print(f"  ✓ {len(_medewerkers)} medewerkers geladen")
        except Exception as e:
            print(f"  ✗ Medewerkers laden mislukt: {e}")

    if os.path.exists(normen_pad):
        try:
            _dienst_normen = laad_dienst_normen(normen_pad)
            print(f"  ✓ {len(_dienst_normen)} dienst normen geladen")
        except Exception as e:
            print(f"  ✗ Normen laden mislukt: {e}")

    server = HTTPServer(('127.0.0.1', poort), Handler)
    url = f'http://127.0.0.1:{poort}'
    print(f"  ✓ App gestart op {url}")
    threading.Timer(1.0, lambda: webbrowser.open(url)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass

if __name__ == '__main__':
    start_server()
