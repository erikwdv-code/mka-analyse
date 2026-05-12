"""
rooster.py — Roostercheck engine voor MKA
Analyseert het rooster op tekorten en genereert ruil/vulsuggesties
"""
import re
from datetime import datetime, timedelta, time


# ── Medewerkersdatabase ────────────────────────────────────────────────────────
# Ingeladen vanuit het Excel bestand, maar hier als fallback/default

MEDEWERKERS = {}   # naam_schoon -> dict met skills etc.
DIENST_NORMEN = {} # dienst_code -> {ma,di,wo,do,vr,za,zo} -> 'groen'|'oranje'|'rood'

# Substituut diensten: als dienst A open staat maar dienst B gevuld is,
# dan is A toch als gedekt te beschouwen
# formaat: {weekdag: {open_dienst: [substituut_diensten]}}
SUBSTITUUT_REGELS = {
    'zondag': {
        '15:30-23:30 C': ['15:30-23:30 B', 'x15:30-23:30 B'],
    },
    'zaterdag': {
        '15:30-23:30 C': ['15:30-23:30 B', 'x15:30-23:30 B'],
    },
}

DAGEN_INDEX = {'maandag': 0, 'dinsdag': 1, 'woensdag': 2,
               'donderdag': 3, 'vrijdag': 4, 'zaterdag': 5, 'zondag': 6}
WEEKDAG_NAMEN = ['maandag','dinsdag','woensdag','donderdag','vrijdag','zaterdag','zondag']


def laad_medewerkers(bestandspad):
    """Laad medewerkers uit Excel bestand."""
    import openpyxl
    wb = openpyxl.load_workbook(bestandspad)
    ws = wb.active

    medewerkers = {}
    for r in range(3, ws.max_row + 1):
        naam = ws.cell(r, 2).value
        if not naam:
            continue
        naam = str(naam).strip()

        functie       = str(ws.cell(r, 3).value or '').strip()
        spoed_b       = str(ws.cell(r, 4).value or '').lower() == 'ja'
        uitgifte_b    = str(ws.cell(r, 5).value or '').lower() == 'ja'
        triage_c_raw  = str(ws.cell(r, 6).value or '').lower()
        triage_c      = triage_c_raw in ('ja', 'inwerk')
        triage_inwerk = triage_c_raw == 'inwerk'
        bp1_a         = str(ws.cell(r, 7).value or '').lower() == 'ja'
        nacht_onthr   = str(ws.cell(r, 8).value or '').lower() == 'ja'
        bijz          = str(ws.cell(r, 9).value or '').strip()

        is_student    = 'student' in functie.lower() or 'student' in bijz.lower()
        # Glaudi: officieel ontheffing maar doet het nog graag
        nacht_soft    = 'graag' in bijz.lower() and 'nachten' in bijz.lower()

        # Voorkeur parsing
        voorkeur_dag     = 'dagdienst' in bijz.lower()
        voorkeur_laat    = 'late' in bijz.lower()
        voorkeur_weekend = 'weekend' in bijz.lower()
        niet_donderdag   = 'niet op donderdag' in bijz.lower()
        # Susan: maand op maand af
        maand_op_af      = 'maand op maand af' in bijz.lower()

        ns = naam_schoon(naam)
        medewerkers[ns] = {
            'naam':           naam,
            'naam_schoon':    ns,
            'functie':        functie,
            'is_student':     is_student,
            'triage_inwerk':  triage_inwerk,  # student, mag niet zelfstandig
            'skills': {
                'spoed_b':    spoed_b,
                'uitgifte_b': uitgifte_b,
                'triage_c':   triage_c,
                'bp1_a':      bp1_a,
            },
            'nacht_ontheffing': nacht_onthr and not nacht_soft,
            'nacht_soft':       nacht_soft,
            'voorkeuren': {
                'dag':        voorkeur_dag,
                'laat':       voorkeur_laat,
                'weekend':    voorkeur_weekend,
                'niet_do':    niet_donderdag,
                'maand_op_af': maand_op_af,
            },
            'bijzonderheden': bijz,
        }

    return medewerkers


def laad_dienst_normen(bestandspad):
    """Laad dienstverdeling normen uit Excel (groen/oranje/rood per dag)."""
    import openpyxl
    wb = openpyxl.load_workbook(bestandspad)
    ws = wb.active

    # Rij 2 = headers: col 3..9 = ma t/m zo
    normen = {}
    for r in range(3, ws.max_row + 1):
        dienst_naam = ws.cell(r, 2).value
        if not dienst_naam:
            continue
        dienst_naam = str(dienst_naam).strip()
        normen[dienst_naam] = {}

        for dag_idx, col in enumerate(range(3, 10)):  # kolommen C t/m I
            cel = ws.cell(r, col)
            val = cel.value
            fill = cel.fill
            bg = fill.fgColor.rgb if fill and fill.fgColor and fill.fgColor.type == 'rgb' else None

            if bg == 'FF92D050':
                status = 'groen'
            elif bg == 'FFFFC000':
                status = 'oranje'
            else:
                status = 'rood'

            dag_naam = WEEKDAG_NAMEN[dag_idx]
            normen[dienst_naam][dag_naam] = {
                'status': status,
                'gevraagd': val == 'x' or status in ('groen', 'oranje'),
            }

    # Normaliseer dienst keys zodat ze matchen met rapport normalisatie
    import re as _re
    genormaliseerd = {}
    for dienst_naam, dag_data in normen.items():
        # Verwijder leading zeros in einduur: 06:00 -> 6:00
        dn = _re.sub(r'-0(\d):', r'-\1:', dienst_naam)
        genormaliseerd[dn] = dag_data
    return genormaliseerd


def naam_schoon(naam):
    if not isinstance(naam, str):
        return ""
    return ' '.join(
        naam.replace('-', ' ')
            .replace('(deta)', '')
            .replace('\xa0', ' ')
            .lower()
            .split()
    )

def extraheer_dienst(dienst_raw):
    """
    Extraheer de relevante dienst uit een mogelijk gecombineerde string.
    '07.00-15.00 A MKA, Ev-AD ARENA' -> '07.00-15.00 A MKA'
    'Stage, x07.00-15.00 C2 MKA' -> 'x07.00-15.00 C2 MKA'
    """
    if not dienst_raw or ',' not in str(dienst_raw):
        return dienst_raw
    delen = [d.strip() for d in str(dienst_raw).split(',')]
    for deel in delen:
        if re.search(r'\d{1,2}[.:]\d{2}', deel):
            return deel
    return delen[0]


def normaliseer_dienst(naam):
    """Normaliseer dienstnaam zodat rapport en normen matchen.
    '07.00-15.00 A MKA' -> '7:00-15:00 A'
    """
    if not naam or naam == '-':
        return None
    d = str(naam).strip()
    d = re.sub(r'\s+MKA\s*$', '', d, flags=re.IGNORECASE).strip()
    d = re.sub(r'(\d{1,2})\.(\d{2})', r'\1:\2', d)
    d = re.sub(r'\b0(\d):', r'\1:', d)
    return d


def zoek_medewerker(naam_in_rooster, medewerkers):
    """Match naam uit rooster naar medewerker in database."""
    ns = naam_schoon(naam_in_rooster)
    if ns in medewerkers:
        return medewerkers[ns]
    # Gedeeltelijke match op achternaam of voornaam
    for key, mw in medewerkers.items():
        delen = ns.split()
        if any(d in key for d in delen if len(d) > 3):
            return mw
    return None


# ── ATW Rusttijden ─────────────────────────────────────────────────────────────

def dienst_tijden(dienst_code):
    """Geef begin- en eindtijd van een dienstcode als time objecten."""
    match = re.search(r'(\d{1,2})[:.:](\d{2})-(\d{1,2})[:.:](\d{2})', dienst_code)
    if not match:
        return None, None
    bh, bm, eh, em = int(match.group(1)), int(match.group(2)), int(match.group(3)), int(match.group(4))
    begin = time(bh, bm)
    eind  = time(eh, em)
    return begin, eind


def eind_datetime(datum, dienst_code):
    """Geef eindtijd als datetime, rekening houdend met diensten die over middernacht gaan."""
    begin, eind = dienst_tijden(dienst_code)
    if begin is None:
        return None
    dt_eind = datetime.combine(datum, eind)
    if eind < begin:  # over middernacht
        dt_eind += timedelta(days=1)
    return dt_eind


def begin_datetime(datum, dienst_code):
    begin, _ = dienst_tijden(dienst_code)
    if begin is None:
        return None
    return datetime.combine(datum, begin)


def atw_ok(vorige_eind_dt, nieuwe_begin_dt, min_rust_uur=11):
    """Check of er voldoende rust is tussen twee diensten (ATW = 11 uur)."""
    if vorige_eind_dt is None or nieuwe_begin_dt is None:
        return True
    rust = (nieuwe_begin_dt - vorige_eind_dt).total_seconds() / 3600
    return rust >= min_rust_uur


def is_nachtdienst(dienst_code):
    """Is dit een nachtdienst (begint 22:00 of later, of eindigt voor 07:00)?"""
    begin, eind = dienst_tijden(dienst_code)
    if begin is None:
        return False
    return begin.hour >= 22 or begin.hour < 6


def is_weekenddienst(datum):
    return datum.weekday() >= 5  # za=5, zo=6


# ── Skill check ────────────────────────────────────────────────────────────────

def dienst_skill(dienst_code):
    """Welke skill is nodig voor deze dienst?"""
    d = dienst_code.upper()
    # Spoed B: begint op 7:00, 15:00 of 23:00 met B
    if re.search(r'(7:00|15:00|23:00).*\bB\b', d) or re.search(r'\b(07:00|15:00|23:00).*\bB\b', d):
        return 'spoed_b'
    # Normale B: 7:30, 15:30
    if re.search(r'(7:30|15:30).*\bB\b', d) or re.search(r'\b(07:30|15:30).*\bB\b', d):
        return 'uitgifte_b'
    # BP1/A
    if d.endswith(' A') or ' A ' in d or d.endswith('\tA'):
        return 'bp1_a'
    # Triage C (alle varianten)
    if re.search(r'\bC\d?\b', d):
        return 'triage_c'
    return None


def kan_dienst_doen(medewerker, dienst_code):
    """Kan deze medewerker deze dienst doen op basis van skills?"""
    if medewerker is None:
        return False, "onbekende medewerker"

    # Studenten mogen niet zelfstandig (triage_inwerk)
    if medewerker['triage_inwerk']:
        return False, "student — mag niet zelfstandig"

    skill = dienst_skill(dienst_code)
    if skill is None:
        return True, ""

    if skill == 'spoed_b' and not medewerker['skills']['spoed_b']:
        return False, "mist spoed-B kwalificatie"
    if skill == 'uitgifte_b' and not medewerker['skills']['uitgifte_b']:
        return False, "mist uitgifte-B kwalificatie"
    if skill == 'bp1_a' and not medewerker['skills']['bp1_a']:
        return False, "mist BP1/senior kwalificatie"
    if skill == 'triage_c' and not medewerker['skills']['triage_c']:
        return False, "mist triage-C kwalificatie"

    # Nachtontheffing
    if is_nachtdienst(dienst_code) and medewerker['nacht_ontheffing']:
        return False, "nachtdienstontheffing"

    return True, ""


def voorkeur_score(medewerker, dienst_code, datum):
    """
    Geef een score hoe goed deze dienst past bij de voorkeuren van de medewerker.
    Hogere score = betere match. Negatief = gaat tegen voorkeur in.
    """
    score = 0
    v = medewerker.get('voorkeuren', {})
    begin, _ = dienst_tijden(dienst_code)
    if begin is None:
        return score

    is_dag  = begin.hour < 12
    is_laat = 14 <= begin.hour < 22
    is_we   = is_weekenddienst(datum)

    if v.get('dag') and is_dag:
        score += 2
    if v.get('dag') and not is_dag:
        score -= 1
    if v.get('laat') and is_laat:
        score += 2
    if v.get('weekend') and is_we:
        score += 1
    if v.get('niet_do') and datum.weekday() == 3:  # donderdag
        score -= 3

    return score


# ── Hoofdanalyse ───────────────────────────────────────────────────────────────

# Substitutieregels: als dienst A open staat maar dienst B gevuld is,
# dan is er geen tekort. Per weekdag gedefinieerd.
# Formaat: {weekdag: {open_dienst: [vervangende_diensten]}}
SUBSTITUUT_REGELS = {
    'zondag': {
        '15:30-23:30 C': ['15:30-23:30 B', 'x15:30-23:30 B'],
    },
}


# Substitutieregels: als een dienst open staat maar er staat een dienst
# met hetzelfde begintijdstip gevuld (B, C, x-variant), dan geen tekort.
SUBSTITUUT_BEGINTIJDEN = {
    'zondag': ['15:30'],
}


def analyseer_rooster(df, medewerkers, dienst_normen):
    """
    Analyseer het rooster per dag:
    - Welke essentiële diensten staan open (niet ingevuld)?
    - Welke medewerkers hebben een rode/niet-gevraagde dienst?
    - Genereer suggesties

    Returns: dict per datum_str -> analyse resultaat
    """
    import pandas as pd

    if df is None or df.empty:
        return {}

    resultaat = {}

    # Groepeer per datum
    for datum, dag_df in df.groupby('Datum'):
        if pd.isna(datum):
            continue

        weekdag = WEEKDAG_NAMEN[datum.weekday()]

        # ── Stap 1: Welke diensten zijn er op deze dag? ───────────────────────
        # Uit het rooster: dienst -> lijst medewerkers
        ingepland = {}   # genormaliseerde dienst_code -> [medewerker_info]
        ingepland_orig = {}  # originele naam voor weergave
        for _, rij in dag_df.iterrows():
            dienst_orig = extraheer_dienst(str(rij.get('Dienst(en) realisatie', '') or '').strip())
            naam        = str(rij.get('Naam', '') or '').strip()
            inzet       = str(rij.get('Inzet', '') or '').strip()

            if not dienst_orig or dienst_orig == '-' or dienst_orig == 'nan':
                continue

            # Inzet 2 = overwerk, maar als er een echte dienst is telt het mee
            import re as _re
            heeft_tijd = bool(_re.search(r'\d{1,2}[.:]\d{2}', dienst_orig))
            heeft_tijd = bool(__import__('re').search(r'\d{1,2}[.:]\d{2}', dienst_orig))
            heeft_tijd = bool(__import__('re').search(r'\d{1,2}[.:]\d{2}', dienst_orig))
            is_overwerk = inzet == 'Inzet 2' and not heeft_tijd and not heeft_tijd and not heeft_tijd
            is_x        = dienst_orig.lower().startswith('x')

            # Normaliseer voor matching met normenbestand
            dienst = normaliseer_dienst(dienst_orig) or dienst_orig
            # x-diensten houden hun x prefix
            if is_x and not dienst.startswith('x'):
                dienst = 'x' + dienst

            mw = zoek_medewerker(naam, medewerkers)

            if dienst not in ingepland:
                ingepland[dienst] = []
                ingepland_orig[dienst] = dienst_orig
            ingepland[dienst].append({
                'naam':       naam,
                'medewerker': mw,
                'is_overwerk': is_overwerk,
                'is_x':        is_x,
            })

        # ── Stap 2: Check normen — open essentiële diensten ──────────────────
        open_groen  = []  # essentieel, niet gevuld
        open_oranje = []  # wenselijk, niet gevuld
        gevuld_rood = []  # rood maar wel gevuld (schuifkandidaten)

        for dienst_naam, norm_dagen in dienst_normen.items():
            norm = norm_dagen.get(weekdag, {})
            status    = norm.get('status', 'rood')
            gevraagd  = norm.get('gevraagd', False)

            # Is de dienst gevuld? (inclusief x-diensten en substituut diensten)
            gevuld = dienst_naam in ingepland and len([
                p for p in ingepland[dienst_naam]
                if not p['is_overwerk']
            ]) > 0

            # Check substituut: is er een dienst met hetzelfde begintijdstip gevuld?
            if not gevuld:
                sub_tijden = SUBSTITUUT_BEGINTIJDEN.get(weekdag, [])
                for sub_tijd in sub_tijden:
                    # Dienst moet BEGINNEN op dat tijdstip (eerste tijdstip in de naam)
                    import re as _re
                    d_begin = _re.match(r'x?(\d{1,2}:\d{2})', dienst_naam)
                    if d_begin and d_begin.group(1) == sub_tijd:
                        # Kijk of ány ingeplande dienst ook op dat tijdstip begint
                        for ing_dienst in ingepland:
                            i_begin = _re.match(r'x?(\d{1,2}:\d{2})', ing_dienst)
                            if i_begin and i_begin.group(1) == sub_tijd and len([
                                p for p in ingepland[ing_dienst] if not p['is_overwerk']
                            ]) > 0:
                                gevuld = True
                                break
                    if gevuld:
                        break

            if not gevuld:
                if status == 'groen':
                    open_groen.append(dienst_naam)
                elif status == 'oranje':
                    open_oranje.append(dienst_naam)
            # Rood maar gevuld = medewerker heeft een niet-essentiële dienst gekozen
            if gevuld and status == 'rood':
                for persoon in ingepland.get(dienst_naam, []):
                    if not persoon['is_x'] and not persoon['is_overwerk']:
                        gevuld_rood.append({
                            'naam':   persoon['naam'],
                            'dienst': dienst_naam,
                            'medewerker': persoon['medewerker'],
                        })

        # ── Stap 3: Suggesties genereren ─────────────────────────────────────
        suggesties = []

        # 3a: Medewerker op rode dienst terwijl groene dienst openstaat
        for rode_inzet in gevuld_rood:
            mw = rode_inzet['medewerker']
            if mw is None:
                continue
            for open_dienst in open_groen:
                kan, reden = kan_dienst_doen(mw, open_dienst)
                if kan:
                    # Check ATW: kijk naar dienst dag ervoor/erna (vereenvoudigd)
                    atw_waarschuwing = _check_atw_context(
                        rode_inzet['naam'], datum, open_dienst, df, medewerkers)
                    suggesties.append({
                        'type':       'wissel_rood_naar_groen',
                        'prioriteit': 1,
                        'medewerker': rode_inzet['naam'],
                        'van_dienst': rode_inzet['dienst'],
                        'naar_dienst': open_dienst,
                        'datum':      datum.strftime('%Y-%m-%d'),
                        'uitleg':     f"{_voornaam(rode_inzet['naam'])} staat op '{rode_inzet['dienst']}' (niet essentieel) — kan verschoven worden naar '{open_dienst}' (essentieel)",
                        'atw_ok':     not atw_waarschuwing,
                        'atw_info':   atw_waarschuwing or '',
                        'voorkeur_score': voorkeur_score(mw, open_dienst, datum),
                    })

        # 3b: Open groene dienst — wie is überhaupt beschikbaar (staat niet al ingepland op die dag)?
        al_ingepland_namen = set()
        for d, personen in ingepland.items():
            for p in personen:
                if not p['is_x']:
                    al_ingepland_namen.add(naam_schoon(p['naam']))

        for open_dienst in open_groen + open_oranje:
            beschikbaar = []
            for ns, mw in medewerkers.items():
                if ns in al_ingepland_namen:
                    continue
                kan, reden = kan_dienst_doen(mw, open_dienst)
                if kan:
                    atw_w = _check_atw_context(mw['naam'], datum, open_dienst, df, medewerkers)
                    score = voorkeur_score(mw, open_dienst, datum)
                    beschikbaar.append({
                        'naam':   mw['naam'],
                        'score':  score,
                        'atw_ok': not atw_w,
                        'atw_info': atw_w or '',
                    })
            # Sorteer: ATW ok eerst, dan hoogste voorkeur score
            beschikbaar.sort(key=lambda x: (0 if x['atw_ok'] else 1, -x['score']))
            suggesties.append({
                'type':        'open_dienst',
                'prioriteit':  1 if open_dienst in open_groen else 2,
                'dienst':      open_dienst,
                'datum':       datum.strftime('%Y-%m-%d'),
                'uitleg':      f"'{open_dienst}' staat open (essentieel)",
                'kandidaten':  beschikbaar[:5],  # top 5
            })

        resultaat[datum.strftime('%Y-%m-%d')] = {
            'datum':       datum.strftime('%Y-%m-%d'),
            'weekdag':     weekdag,
            'ingepland':   {k: [{'naam': p['naam'], 'is_x': p['is_x']} for p in v]
                           for k, v in ingepland.items()},
            'open_groen':  open_groen,
            'open_oranje': open_oranje,
            'gevuld_rood': gevuld_rood,
            'suggesties':  suggesties,
            'tekort':      len(open_groen) + len(open_oranje),
        }

    return resultaat


def _voornaam(naam):
    """Geef voornaam van een naam (laatste woord in 'Achternaam Voornaam' formaat)."""
    delen = naam.strip().split()
    return delen[-1] if delen else naam


def _check_atw_context(naam, datum, nieuwe_dienst, df, medewerkers):
    """
    Check ATW voor een medewerker: kijk naar de dag voor en na.
    Geeft een waarschuwingsstring terug als ATW niet ok is, anders None.
    """
    import pandas as pd
    ns = naam_schoon(naam)
    dag_voor  = datum - timedelta(days=1)
    dag_na    = datum + timedelta(days=1)

    nieuwe_begin = begin_datetime(datum, nieuwe_dienst)
    nieuwe_eind  = eind_datetime(datum, nieuwe_dienst)

    # Dienst dag ervoor
    df_voor = df[df['Datum'] == pd.Timestamp(dag_voor)]
    for _, rij in df_voor.iterrows():
        if naam_schoon(str(rij.get('Naam', ''))) == ns:
            vorige_dienst = str(rij.get('Dienst(en) realisatie', '') or '')
            if vorige_dienst and vorige_dienst != '-':
                vorige_eind = eind_datetime(dag_voor, vorige_dienst)
                if not atw_ok(vorige_eind, nieuwe_begin):
                    rust = (nieuwe_begin - vorige_eind).total_seconds() / 3600 if vorige_eind and nieuwe_begin else 0
                    return f"⚠ ATW: slechts {rust:.0f}u rust na dienst op {dag_voor.strftime('%d-%m')}"

    # Dienst dag erna
    df_na = df[df['Datum'] == pd.Timestamp(dag_na)]
    for _, rij in df_na.iterrows():
        if naam_schoon(str(rij.get('Naam', ''))) == ns:
            volgende_dienst = str(rij.get('Dienst(en) realisatie', '') or '')
            if volgende_dienst and volgende_dienst != '-':
                volgende_begin = begin_datetime(dag_na, volgende_dienst)
                if not atw_ok(nieuwe_eind, volgende_begin):
                    rust = (volgende_begin - nieuwe_eind).total_seconds() / 3600 if nieuwe_eind and volgende_begin else 0
                    return f"⚠ ATW: slechts {rust:.0f}u rust voor dienst op {dag_na.strftime('%d-%m')}"

    return None
