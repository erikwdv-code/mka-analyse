import pandas as pd
import re
from datetime import datetime
from rooster import extraheer_dienst


SENIOREN = [
    "harten anton", "wijker bianca", "timmer ellen", "sant gert",
    "kruisbergen hans", "bulsing catsburg ida", "moorten ilonka",
    "sluijs ina", "groot muller kimberly", "nikken laura",
    "snoek leontien", "glim bloedjes linda", "alblas marjan",
    "zandwijk mark", "kleinepier rianne"
]

# ── Categorie definities ───────────────────────────────────────────────────────
# Indirect productief: exacte dienstcodes (hoofdletterongevoelig)
INDIRECT_CODES = {
    's9 locatie', 's', 'plb dienst', 'or8-scholing', 'or8', 'les/mka',
    'kck', 'k9', 'k8 mka', 'k8', 'brain', 'ava bilthoven',
    'a2 brede triage', 'werkoverleg', 'vakantie dienst',
    'k4 twvg', 'k4', 'ev-ad arena', 'stage',
}

# Absent: exacte dienstcodes (hoofdletterongevoelig)
ABSENT_CODES = {
    'tvt opname d', 'tvt opname c', 'senior8',
    'feestdagen verlof', 'compensatieverlof dienst',
    'compensatieverlof contract', 'ziek dienst <24 u',
    'ziek dienst', 'ziek contract',
}


def naam_schoon(naam):
    if not isinstance(naam, str):
        return ""
    return ' '.join(naam.replace('-', ' ').replace('(deta)', '').replace('\xa0', ' ').lower().split())


def laad_uitsluitingen(bestandspad):
    """Laad lijst van uitgesloten medewerkers."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(bestandspad)
        ws = wb.active
        namen = set()
        for r in range(2, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val:
                namen.add(naam_schoon(str(val)))
        return namen
    except Exception:
        return set()


def is_uitgesloten(naam, uitsluitingen):
    ns = naam_schoon(naam)
    return ns in uitsluitingen


def is_senior(naam):
    return naam_schoon(naam) in SENIOREN


def heeft_tijdpatroon(dienst):
    if not isinstance(dienst, str):
        return False
    return bool(re.search(r'\d{1,2}[.:]\d{2}', dienst))


def classificeer_dienst(rij, uitsluitingen=None):
    naam    = str(rij.get('Naam', '') or '')
    functie = str(rij.get('Hoofdfunctie', '') or '')
    inzet   = str(rij.get('Inzet', '') or '')
    dienst  = str(rij.get('Dienst(en) realisatie', '') or '')
    dienst_lower = dienst.lower().strip()

    # Uitgesloten medewerkers
    if uitsluitingen and is_uitgesloten(naam, uitsluitingen):
        return 'uitsluiten'

    # Studenten uitsluiten
    if 'student' in functie.lower():
        if dienst_lower == 'les/mka':
            return 'uitsluiten'

    # Overwerk: alleen als er geen echte MKA dienst (tijdpatroon) in zit
    if inzet == 'Inzet 2' and not heeft_tijdpatroon(dienst):
        return 'overwerk'

    # Lege of streepje
    if not dienst or dienst == '-' or dienst.strip() == '':
        return 'uitsluiten'

    # X-diensten: begint met x én heeft tijdpatroon
    if dienst_lower.startswith('x') and heeft_tijdpatroon(dienst):
        return 'x-diensten'

    # Absent: exacte match
    if dienst_lower in ABSENT_CODES:
        return 'absent'

    # Indirect productief: exacte match
    if dienst_lower in INDIRECT_CODES:
        return 'indirect'

    # Productief: heeft tijdpatroon
    if heeft_tijdpatroon(dienst):
        return 'productief'

    # Overig vangnet → indirect
    return 'indirect'


def laad_rapport(bestandspad):
    """Laad een exportbestand en geef een DataFrame terug."""
    try:
        df = pd.read_excel(bestandspad, sheet_name=0, header=0)
    except Exception as e:
        raise ValueError(f"Kan bestand niet openen: {e}")

    verwacht = ['Datum', 'Naam', 'Hoofdfunctie', 'Inzet', 'Dienst(en) realisatie']
    aanwezig = [k for k in verwacht if k in df.columns]
    if len(aanwezig) < 4:
        raise ValueError(f"Bestand mist verwachte kolommen. Gevonden: {list(df.columns)}")

    if 'Datum' in df.columns:
        df['Datum'] = pd.to_datetime(df['Datum'], errors='coerce')

    df['naam_schoon'] = df['Naam'].apply(naam_schoon)
    df['IsSenior'] = df['naam_schoon'].isin(SENIOREN)

    return df


def bereken_dienstverdeling(df, uitsluitingen=None):
    """Bereken verdeling productief / indirect / absent / x-diensten / overwerk."""
    df = df.copy()
    df['categorie'] = df.apply(
        lambda r: classificeer_dienst(r.to_dict(), uitsluitingen), axis=1)

    telling = df['categorie'].value_counts().to_dict()
    productief = telling.get('productief', 0)
    indirect   = telling.get('indirect', 0)
    absent     = telling.get('absent', 0)
    xdiensten  = telling.get('x-diensten', 0)
    overwerk   = telling.get('overwerk', 0)
    totaal     = productief + indirect + absent + xdiensten

    def pct(n):
        return round(n / totaal * 100, 1) if totaal > 0 else 0

    _DAGNAMEN = ['maandag','dinsdag','woensdag','donderdag','vrijdag','zaterdag','zondag']
    _MAANDEN  = ['','januari','februari','maart','april','mei','juni',
                 'juli','augustus','september','oktober','november','december']

    def detail_rijen(cat):
        sub = df[df['categorie'] == cat].copy()
        rijen = []
        for _, r in sub.iterrows():
            datum = r['Datum']
            if hasattr(datum, 'strftime'):
                datum_str = datum.strftime('%Y-%m-%d')
                datum_nl  = f"{_DAGNAMEN[datum.weekday()]} {datum.day} {_MAANDEN[datum.month]} {datum.year}"
            else:
                datum_str = str(datum)
                datum_nl  = datum_str
            rijen.append({
                'datum_str': datum_str,
                'datum_nl':  datum_nl,
                'naam':      str(r.get('Naam', '')),
                'dienst':    str(r.get('Dienst(en) realisatie', '')),
            })
        rijen.sort(key=lambda x: (x['datum_str'], x['naam']))
        return rijen

    return {
        'productief': {'aantal': productief, 'pct': pct(productief), 'detail': detail_rijen('productief')},
        'indirect':   {'aantal': indirect,   'pct': pct(indirect),   'detail': detail_rijen('indirect')},
        'absent':     {'aantal': absent,      'pct': pct(absent),     'detail': detail_rijen('absent')},
        'xdiensten':  {'aantal': xdiensten,  'pct': pct(xdiensten),  'detail': detail_rijen('x-diensten')},
        'overwerk':   {'aantal': overwerk,    'pct': None,            'detail': detail_rijen('overwerk')},
        'totaal':     totaal,
        'df':         df,
    }


def bereken_senioren_overzicht(df):
    """Per dag: welke senioren staan ingeroosterd, op welke dienst."""
    df_senior = df[df['IsSenior'] == True].copy()
    df_senior = df_senior[df_senior['Datum'].notna()]
    df_senior = df_senior[df_senior['Dienst(en) realisatie'].notna()]
    df_senior = df_senior[df_senior['Dienst(en) realisatie'] != '-']

    if df_senior.empty:
        return {}

    def voornaam(naam):
        s = naam_schoon(naam)
        delen = s.split()
        return delen[-1].capitalize() if delen else naam

    df_senior['voornaam'] = df_senior['Naam'].apply(voornaam)

    overzicht = {}
    for datum, groep in df_senior.groupby('Datum'):
        datum_str = datum.strftime('%Y-%m-%d')
        dag_data = []
        for _, rij in groep.iterrows():
            dag_data.append({
                'naam':   rij['voornaam'],
                'dienst': str(rij['Dienst(en) realisatie']),
            })
        dag_data.sort(key=lambda x: x['dienst'])
        overzicht[datum_str] = dag_data

    return overzicht


def vergelijk_rapporten(df1, label1, df2, label2, uitsluitingen=None):
    v1 = bereken_dienstverdeling(df1, uitsluitingen)
    v2 = bereken_dienstverdeling(df2, uitsluitingen)
    return {'label1': label1, 'label2': label2, 'verdeling1': v1, 'verdeling2': v2}
